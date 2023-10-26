import hashlib
import random
import threading
import timeit

import openpyxl
import pandas as pd

import HashingAndSalting

'''
Euclid's algorithm for determining the greatest common divisor
Use iteration to make it faster for larger integers
'''


def gcd(a, b):
    while b != 0:
        a, b = b, a % b
    return a


'''
Tests to see if a number is prime.
'''


def is_prime(num):
    if num == 2:
        return True
    if num < 2 or num % 2 == 0:
        return False
    for n in range(3, int(num ** 0.5) + 2, 2):
        if num % n == 0:
            return False
    return True


'''
Euclid's extended algorithm for finding the multiplicative inverse of two numbers
'''


def multiplicative_inverse(e, phi):
    d = 0
    x1 = 0
    x2 = 1
    y1 = 1
    temp_phi = phi

    while e > 0:
        temp1 = (int)(temp_phi / e)
        temp2 = temp_phi - temp1 * e
        temp_phi = e
        e = temp2

        x = x2 - temp1 * x1
        y = d - temp1 * y1

        x2 = x1
        x1 = x
        d = y1
        y1 = y

    if temp_phi == 1:
        return d + phi


def generate_keypair(p, q, r):
    if not (is_prime(p) and is_prime(q)):
        raise ValueError('Both numbers must be prime.')
    elif p == q:
        raise ValueError('p and q cannot be equal')
    # n = pqr
    n = p * q * r

    # Phi is the totient of n
    phi = (p - 1) * (q - 1) * (r - 1)

    # Choose an integer e such that e and phi(n) are coprime
    e = random.randrange(1, phi)

    # Use Euclid's Algorithm to verify that e and phi(n) are comprime
    g = gcd(e, phi)

    while g != 1:
        e = random.randrange(1, phi)
        g = gcd(e, phi)

        # Use Extended Euclid's Algorithm to generate the private key
    d = multiplicative_inverse(e, phi)

    # Return public and private keypair
    # Public key is (e, n) and private key is (d, n)
    return ((e, n), (d, n), phi)


def encrypt(plaintext,nn):
    data1 = pd.read_csv("data1.csv")
    data2 = pd.read_csv("data2.csv")

    key = int(data2['0'][1])  # e
    n = int(data1['0'][3])

    wb = openpyxl.load_workbook("storedhashandsalt.xlsx")
    sh1 = wb['hashed']
    sh1.cell(row=nn, column=5, value=key)
    wb.save("storedhashandsalt.xlsx")

    # Convert each letter in the plaintext to numbers based on the character using a^b mod m
    cipher = [(ord(char) ** int(key)) % n for char in plaintext]
    # Return the array of bytes
    return cipher


def decrypt(ciphertext):
    data1 = pd.read_csv("data1.csv")
    data2 = pd.read_csv("data2.csv")

    key = int(data2['0'][2])  # d
    n = int(data1['0'][3])

    # Generate the plaintext based on the ciphertext and key using a^b mod m
    plain = [chr((char ** key) % n) for char in ciphertext]
    # Return the array of bytes as a string
    return ''.join(plain)


def callme(message,nn):

    primes = []
    total_no_primes = 0
    with open('primes.txt') as pfile:
        for line in pfile:
            primes.append(int(line))  # = [int(i) for i in line.split()]
            total_no_primes += 1
    p = primes[random.randint(1, total_no_primes - 1)]
    q = primes[random.randint(1, total_no_primes - 1)]
    r = primes[random.randint(1, total_no_primes - 1)]

    print("Generating your public/private keypairs now . . .")
    public, private, phi = generate_keypair(p, q, r)
    print("\nYour public key is ", public, " and your private key is ", private)

    data1 = [p, q, phi, public[1]]  # p, q, phi, n
    df = pd.DataFrame(data1)
    df.to_csv('data1.csv')  # offline storage of p, q, phi, n in table 1

    data2 = [r, public[0], private[0]]  # r, e, d
    df = pd.DataFrame(data2)
    df.to_csv('data2.csv')  # offline storage of r, e, d in table 2

    tic = timeit.default_timer()
    encrypted_msg = encrypt(message,nn)
    print("\nYour encrypted message is: ")
    print(''.join([str(x) for x in encrypted_msg]))
    toc = timeit.default_timer()

    # encrypted_salted_text = HashingAndSalting.salt(encrypted_msg)
    # Sender_HASH = hashlib.sha256(encrypted_salted_text.encode()).hexdigest()
    # print(Sender_HASH)
    wb = openpyxl.load_workbook("storedhashandsalt.xlsx")
    sh1 = wb['hashed']
    sh1.cell(row=threading.active_count(), column=2, value=''.join([str(x) for x in encrypted_msg]))
    sh1.cell(row=threading.active_count(), column=1, value='192.168.56.1')
    sh1.cell(row=threading.active_count(), column=3, value=5050)
    # sh1.cell(row=threading.activeCount(), column=4,value=timeit.default_timer())
    sh1.cell(row=threading.active_count(), column=4, value=toc - tic)

    # sh1.cell(row=threading.active_count(), column=5, value=nm)



    wb.save("storedhashandsalt.xlsx")
    print("\nDecrypting message with private key ", private, " . . .")
    print("\nYour decrypt message is:")
    print(decrypt(encrypted_msg))